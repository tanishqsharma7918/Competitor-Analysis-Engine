from typing import Dict, Any, List
import pandas as pd
from io import BytesIO
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


def generate_excel_report(
    product_name: str,
    competitors: List[Dict[str, Any]],
    matrix_df: pd.DataFrame,
    analysis: Dict[str, Any],
    features: List[Dict[str, Any]]
) -> bytes:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    ws_summary.append([f"Competitor Analysis Report: {product_name}"])
    ws_summary["A1"].font = Font(bold=True, size=16)
    ws_summary.append([])
    
    ws_summary.append(["Section", "Details"])
    ws_summary["A3"].fill = header_fill
    ws_summary["A3"].font = header_font
    ws_summary["B3"].fill = header_fill
    ws_summary["B3"].font = header_font
    
    ws_summary.append(["Product Name", product_name])
    ws_summary.append(["Number of Competitors", len(competitors)])
    ws_summary.append(["Number of Features Analyzed", len(features)])
    
    ws_competitors = wb.create_sheet("Competitors")
    ws_competitors.append(["Company Name", "Product Name", "Description", "Website", "Market Position"])
    for cell in ws_competitors[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    for comp in competitors:
        ws_competitors.append([
            comp.get("company_name", ""),
            comp.get("product_name", ""),
            comp.get("description", ""),
            comp.get("website", ""),
            comp.get("market_position", "")
        ])
    
    for col in ws_competitors.columns:
        for cell in col:
            cell.border = border
    
    ws_matrix = wb.create_sheet("Feature Matrix")
    for r in dataframe_to_rows(matrix_df, index=False, header=True):
        ws_matrix.append(r)
    
    for cell in ws_matrix[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    for row in ws_matrix.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            if cell.value == "✓":
                cell.fill = green_fill
            elif cell.value == "✗":
                cell.fill = red_fill
    
    ws_diff = wb.create_sheet("Differentiators")
    ws_diff.append(["Type", "Title", "Description"])
    for cell in ws_diff[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    for diff in analysis.get("differentiators", []):
        ws_diff.append(["Differentiator", diff.get("title", ""), diff.get("description", "")])
    
    for rec in analysis.get("recommendations", []):
        ws_diff.append(["Recommendation", rec.get("title", ""), rec.get("description", "")])
    
    for miss in analysis.get("missing_capabilities", []):
        ws_diff.append([
            "Missing Capability",
            f"{miss.get('capability', '')} ({miss.get('importance', '')})",
            miss.get('rationale', '')
        ])
    
    for col in ws_diff.columns:
        for cell in col:
            cell.border = border
    
    for column in ws_summary.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min((max_length + 2), 50)
        ws_summary.column_dimensions[column[0].column_letter].width = adjusted_width
    
    for ws in [ws_competitors, ws_matrix, ws_diff]:
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min((max_length + 2), 50)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generate_pdf_report(
    product_name: str,
    competitors: List[Dict[str, Any]],
    matrix_df: pd.DataFrame,
    analysis: Dict[str, Any],
    features: List[Dict[str, Any]]
) -> bytes:

    # ---------- CLEAN & SANITIZE TEXT ----------
    def clean(text: str) -> str:
        """Convert to ASCII-only and remove unsupported characters."""
        if not isinstance(text, str):
            text = str(text)
        ascii_text = ''.join(c for c in text if ord(c) < 128)
        return break_long_words(ascii_text)

    def break_long_words(text: str, max_len: int = 45) -> str:
        """Prevent long tokens from breaking PDF width."""
        words = text.split()
        out = []
        for w in words:
            if len(w) > max_len:
                out.append(w[:max_len] + "-")
                out.append(w[max_len:])
            else:
                out.append(w)
        return " ".join(out)

    # ---------- INIT PDF ----------
    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    full_width = pdf.w - (pdf.l_margin + pdf.r_margin)  # safe usable width

    # ---------- TITLE ----------
    pdf.set_font("Helvetica", "B", 20)
    pdf.multi_cell(full_width, 10, clean("Competitor Analysis Report"), align="C")
    pdf.ln(5)

    # ---------- PRODUCT ----------
    pdf.set_font("Helvetica", "B", 14)
    pdf.multi_cell(full_width, 8, clean(f"Product: {product_name}"))
    pdf.ln(5)

    # ---------- EXEC SUMMARY ----------
    pdf.set_font("Helvetica", "B", 12)
    pdf.multi_cell(full_width, 8, clean("Executive Summary"))

    pdf.set_font("Helvetica", "", 10)
    summary = (
        f"This report analyzes {product_name} against "
        f"{len(competitors)} competitors across {len(features)} key features."
    )
    pdf.multi_cell(full_width, 6, clean(summary))
    pdf.ln(5)

    # ---------- COMPETITORS ----------
    pdf.set_font("Helvetica", "B", 12)
    pdf.multi_cell(full_width, 8, clean("Competitors Identified"))

    pdf.set_font("Helvetica", "", 10)
    for i, comp in enumerate(competitors, start=1):
        pdf.multi_cell(full_width, 6, clean(
            f"{i}. {comp.get('company_name','Unknown')} - {comp.get('product_name','Unknown')}"
        ))

        desc = comp.get("description", "")
        if desc:
            pdf.set_font("Helvetica", "I", 9)
            pdf.multi_cell(full_width, 5, clean("  " + desc))
            pdf.set_font("Helvetica", "", 10)

    pdf.ln(5)

    # ---------- DIFFERENTIATORS ----------
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 12)
    pdf.multi_cell(full_width, 8, clean("Key Differentiators"))

    pdf.set_font("Helvetica", "", 10)
    for i, diff in enumerate(analysis.get("differentiators", []), start=1):
        pdf.set_font("Helvetica", "B", 10)
        pdf.multi_cell(full_width, 6, clean(f"{i}. {diff.get('title','')}"))
        pdf.set_font("Helvetica", "", 10)
        pdf.multi_cell(full_width, 6, clean("  " + diff.get("description", "")))
        pdf.ln(2)

    # ---------- RECOMMENDATIONS ----------
    pdf.ln(5)
    pdf.set_font("Helvetica", "B", 12)
    pdf.multi_cell(full_width, 8, clean("Strategic Recommendations"))

    pdf.set_font("Helvetica", "", 10)
    for i, rec in enumerate(analysis.get("recommendations", []), start=1):
        pdf.set_font("Helvetica", "B", 10)
        pdf.multi_cell(full_width, 6, clean(f"{i}. {rec.get('title','')}"))
        pdf.set_font("Helvetica", "", 10)
        pdf.multi_cell(full_width, 6, clean("  " + rec.get("description", "")))
        pdf.ln(2)

    # ---------- MISSING CAPABILITIES (5 ONLY) ----------
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 12)
    pdf.multi_cell(full_width, 8, clean("Missing Capabilities"))

    pdf.set_font("Helvetica", "", 10)
    for i, miss in enumerate(analysis.get("missing_capabilities", [])[:5], start=1):
        pdf.set_font("Helvetica", "B", 10)
        pdf.multi_cell(
            full_width,
            6,
            clean(f"{i}. {miss.get('capability','')} (Importance: {miss.get('importance','Unknown')})")
        )
        pdf.set_font("Helvetica", "", 10)
        pdf.multi_cell(full_width, 6, clean("  " + miss.get("rationale", "")))
        pdf.ln(2)

    # ---------- FEATURE SUMMARY (7 ONLY) ----------
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 12)
    pdf.multi_cell(full_width, 8, clean("Feature Comparison Summary"))

    pdf.set_font("Helvetica", "", 9)
    pdf.multi_cell(full_width, 6, clean(
        "Note: See Excel report for complete feature matrix (PDF space is limited)."
    ))
    pdf.ln(3)

    for i, feature in enumerate(features[:7], start=1):
        pdf.multi_cell(
            full_width,
            5,
            clean(f"{i}. {feature.get('feature_name','')} ({feature.get('category','General')})")
        )

    pdf.ln(4)

    # ---------- OUTPUT PDF ----------
    return pdf.output(dest="S").encode("latin-1")
